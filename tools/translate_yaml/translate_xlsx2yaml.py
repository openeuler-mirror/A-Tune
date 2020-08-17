#!/usr/bin/python3
# -*- coding: utf-8 -*-
# Copyright (c) 2020 Huawei Technologies Co., Ltd.
# A-Tune is licensed under the Mulan PSL v2.
# You can use this software according to the terms and conditions of the Mulan PSL v2.
# You may obtain a copy of Mulan PSL v2 at:
#     http://license.coscl.org.cn/MulanPSL2
# THIS SOFTWARE IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND, EITHER EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT, MERCHANTABILITY OR FIT FOR A PARTICULAR
# PURPOSE.
# See the Mulan PSL v2 for more details.
# Create: 2020-07-30

"""
translate .xlsx file to .yaml files
"""

import string
import io

import openpyxl
from translate_yaml import TranslateYaml


class TranslateXlsx2Yaml(TranslateYaml):
    """
    The subclass for translating xlsx to .yaml files
    """

    def __init__(self, in_file_name, out_file_name, project_name,
                 iterations, block_dev, network_dev, test):
        super(TranslateXlsx2Yaml, self).__init__(out_file_name, project_name,
                                                 iterations, block_dev, network_dev, test)
        with open(in_file_name, 'rb') as file:
            in_mem_file = io.BytesIO(file.read())
        self.workbook = openpyxl.load_workbook(in_mem_file, read_only=True)

    @staticmethod
    def read_line(reader, line):
        """
        Get a line from xlsx into list.
        :param reader: the content of the xlsx to be read.
        :param line: the number of line to get.
        :return: list, the yaml object.
        """
        obj_list = []
        name_ = reader[str('B' + str(line))].value
        if name_ is None or name_.strip() == "":
            return obj_list

        cols = string.ascii_uppercase[1:15]
        for col in cols:
            val = reader[str(col + str(line))].value
            if val is None:
                val = ''
            obj_list.append(str(val))

        return obj_list

    def translate(self):
        """
        Translate xlsx to yaml.
        :return: True or False, translation result.
        """
        sheetnames = self.workbook.sheetnames
        worksheet = self.workbook[sheetnames[0]]
        line = 2
        return self.write_yaml(worksheet, line)
