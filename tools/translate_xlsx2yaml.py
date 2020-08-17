#!/usr/bin/python3
# -*- coding: utf-8 -*-
# Copyright (c) 2020 Huawei Technologies Co., Ltd.
# A-Tune is licensed under the Mulan PSL v2.
# You can use this software according to the terms and conditions of the Mulan PSL v2.
# You may obtain a copy of Mulan PSL v2 at:
#     http://license.coscl.org.cn/MulanPSL2
# THIS SOFTWARE IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND, EITHER EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT, MERCHANTABILITY OR FIT FOR A PARTICULAR
# PURPOSE.
# See the Mulan PSL v2 for more details.
# Create: 2020-07-30

"""
The tool to translate .excel file to .yaml files
Usage: python3 translate_xlsx2yaml.py [-h] [-i] [-o] [-t] [-p]
"""

import argparse
import os
import subprocess
import string
from enum import Enum
from io import StringIO
import io
import openpyxl
import glob
import shlex


class CsvAttr(Enum):
    """Enumerate the headers"""
    name = 0
    desc = 1
    geti = 2
    seti = 3
    needrestart = 4
    typei = 5
    options = 6
    dtype = 7
    scope_min = 8
    scope_max = 9
    step = 10
    items = 11
    select = 12


class TuningObject:
    """Tuning Object"""
    def __init__(self, obj_list, block_dev, network_dev):
        self.name = obj_list[CsvAttr.name.value].strip()
        self.desc = obj_list[CsvAttr.desc.value].strip()
        self.get = obj_list[CsvAttr.geti.value].strip()
        self.set = obj_list[CsvAttr.seti.value].strip()
        self.needrestart = obj_list[CsvAttr.needrestart.value].strip()
        self.type = obj_list[CsvAttr.typei.value].strip()
        self.options = obj_list[CsvAttr.options.value].strip()
        self.dtype = obj_list[CsvAttr.dtype.value].strip()
        self.scope_min = obj_list[CsvAttr.scope_min.value].strip()
        self.scope_max = obj_list[CsvAttr.scope_max.value].strip()
        self.step = obj_list[CsvAttr.step.value].strip()
        self.items = obj_list[CsvAttr.items.value].strip()
        self.block_device = block_dev
        self.network_device = network_dev

        if '@name' in self.get:
            self.get = self.get.replace('@name', self.name)
        if '@name' in self.set:
            self.set = self.set.replace('@name', self.name)

        if '@block' in self.get:
            self.get = self.get.replace('@block', self.block_device)
        if '@block' in self.set:
            self.set = self.set.replace('@block', self.block_device)

        if '@netdev' in self.get:
            self.get = self.get.replace('@netdev', self.network_device)
        if '@netdev' in self.set:
            self.set = self.set.replace('@netdev', self.network_device)

        self.needrestart = self.needrestart.lower()

    def exec_cmd(self, cmd, value):
        """
        Execute the input command
        :param cmd: input command
        :param value: the value used in the command
        :return: True or False, the results of the execution
        """

        if 'echo' in self.set:
            cmd1, cmd2 = cmd.split('>')
            try:
                file = open(cmd2.strip(), "w")
            except PermissionError:
                return False
            process = subprocess.Popen(shlex.split(cmd1), stdout=file, shell=False)
            process.wait()
            result = str(process.returncode)
        elif 'ifconfig' in self.set:
            process = subprocess.Popen(shlex.split(cmd), shell=False)
            process.wait()
            result = str(process.returncode)
        elif 'sysctl' in self.set:
            result = subprocess.check_output(shlex.split(cmd), stderr=subprocess.STDOUT, shell=False).decode().strip()
        else:
            process = subprocess.Popen(shlex.split(cmd), shell=False)
            process.wait()
            result = str(process.returncode)

        cmd_fail = False
        if 'sysctl' in self.set:
            if 'Invalid argument' in result:
                cmd_fail = True
        elif 'ethtool' in self.set:
            if result != '0' and result != '80':
                cmd_fail = True
        else:
            if result != '0':
                cmd_fail = True

        if cmd_fail:
            print('Invalid Value for ', self.name, ' :', value, '. Make sure the value is in valid range!(', cmd, ')')
            return False

        return True

    def valid_test_command(self, value, orig_value):
        """
        Test the set command in excel
        :param value: the value used in the set command
        :param orig_value: the result of get command
        :return: True or False, the result of execution set command
        """
        if value == orig_value:
            return True

        cmd = self.set.replace('$value', value)
        result = self.exec_cmd(cmd, value)
        return result


    def valid_test_good(self):
        """
        Main check, check the command and given values.
        :return: True or False, check result
        """
        if self.type == 'continuous':
            return True

        get_cmd = self.get
        count = get_cmd.count('|')
        if count > 0:
            get_cmds = get_cmd.split('|')
            i = 0
            for cmds in get_cmds:
                if i == 0:
                    child = subprocess.Popen(shlex.split(cmds), shell=False, stdout=subprocess.PIPE)
                elif i == count:
                    process = subprocess.Popen(shlex.split(cmds), stdin=child.stdout,
                                               stdout=subprocess.PIPE, shell=False)
                    result = process.communicate()
                else:
                    child = subprocess.Popen(shlex.split(cmds), shell=False, stdin=child.stdout, stdout=subprocess.PIPE)
                i += 1
        else:
            process = subprocess.Popen(shlex.split(get_cmd), shell=False, stdout=subprocess.PIPE)
            result = process.communicate()

        orig_value = str(result)

        if self.dtype == 'int':
            result_min = self.valid_test_command(self.scope_min, orig_value)
            result_max = self.valid_test_command(self.scope_max, orig_value)
            if not result_min or not result_max:
                return False
        elif self.dtype == 'string':
            dis_options = self.options.split(';')
            for option in dis_options:
                option = '\"' + option.strip() + '\"'
                result = self.valid_test_command(option, orig_value)
                if not result:
                    return False
        else:
            print("Invalid dtype of ", self.name, ":", self.dtype, ". Only support int and string")
            return False

        return True

    def __repr__(self):
        fstr = StringIO()
        fstr.write('  -\n')
        fstr.write('    name : \"' + self.name + '\"\n')
        fstr.write('    info :\n')
        fstr.write('        desc : \"' + self.desc + '\"\n')
        fstr.write('        get : \"' + self.get + '\"\n')
        fstr.write('        set : \"' + self.set + '\"\n')
        fstr.write('        needrestart : \"' + self.needrestart + '\"\n')
        fstr.write('        type : \"' + self.type + '\"\n')

        if self.type == 'continuous':
            fstr.write('        scope :\n')
            fstr.write('          - ' + self.scope_min + '\n')
            fstr.write('          - ' + self.scope_max + '\n')
            if self.dtype == 'string':
                fstr.write('        dtype : \"string\"\n')
            else:
                fstr.write('        dtype : \"int\"\n')
        elif self.type == 'discrete':
            if self.dtype == 'string':
                fstr.write('        options :\n')
                dis_options = self.options.split(';')
                for option in dis_options:
                    fstr.write('          - \"' + option.strip() + '\"\n')
                fstr.write('        dtype : \"string\"\n')
            elif self.dtype == 'int':
                fstr.write('        scope :\n')
                fstr.write('          - ' + self.scope_min + '\n')
                fstr.write('          - ' + self.scope_max + '\n')
                fstr.write('        step : ' + self.step + '\n')
                fstr.write('        items : \n')
                if self.items != '':
                    dis_items = self.items.split(';')
                    for dit in dis_items:
                        fstr.write('          - ' + dit + '\n')
                fstr.write('        dtype : \"int\"\n')
            else:
                pass
        else:
            pass

        return fstr.getvalue()


class XLSX2YAML:
    """
    Get objects from excel files
    """
    def __init__(self, in_file_name, out_file_name, project_name, iterations):
        with open(in_file_name, 'rb') as file:
            in_mem_file = io.BytesIO(file.read())
        self.workbook = openpyxl.load_workbook(in_mem_file, read_only=True)
        self.out_file = open(out_file_name, 'w')
        self.project_name = project_name
        self.iterations = iterations

    def __del__(self):
        self.out_file.close()

    def get_head(self):
        """
        Generate the header of yaml file.
        :return: string, the header of yaml file.
        """
        fstr = StringIO()
        fstr.write('project:' + ' \"' + self.project_name + '\"' + '\n')
        fstr.write('maxiterations: ' + str(self.iterations) + '\n')
        fstr.write('startworkload: \"\"\n')
        fstr.write('stopworkload: \"\"\n')
        fstr.write('object : \n')
        return fstr.getvalue()

    def read_line_tuning_object(self, worksheet, line):
        """
        Get a line of excel into list.
        :param worksheet: excel sheet.
        :param line: the number of line to get.
        :return: list, the tuning object.
        """
        obj_list = []
        name_ = worksheet[str('B' + str(line))].value
        if name_ is None or name_.strip() == "":
            return obj_list

        cols = string.ascii_uppercase[1:15]
        for col in cols:
            val = worksheet[str(col + str(line))].value
            if val is None:
                val = ''
            obj_list.append(str(val))

        return obj_list

    def translate(self, block_dev, network_dev, test):
        """
        Translate excel to yaml.
        :param block_dev: the name of block device.
        :param network_dev: the name of network device.
        :param test: whether test the commands in file or not.
        :return: True or False, translation result.
        """
        self.out_file.write(self.get_head())
        sheetnames = self.workbook.sheetnames
        worksheet = self.workbook[sheetnames[0]]
        line = 2
        obj_list = self.read_line_tuning_object(worksheet, line)
        if not obj_list:
            print("Empty workbook, translation stops:", self.out_file)
            return False

        while obj_list:
            is_select = obj_list[CsvAttr.select.value].strip()
            if is_select == 'yes':
                tun_obj = TuningObject(obj_list, block_dev, network_dev)
                if test == 'True':
                    valid = tun_obj.valid_test_good()
                    if not valid:
                        return False

                if tun_obj.name != '':
                    self.out_file.write(str(tun_obj))
            line = line + 1
            obj_list = self.read_line_tuning_object(worksheet, line)

        return True


def main(in_dir, out_dir, iterations, project_name, block_dev, network_dev, test):
    """
    Translate .xlsx files to .yaml files.
    :param in_dir: the folder of input excel files.
    :param out_dir: the folder of output yaml files.
    :param iterations: iterations of the project (> 10).
    :param project_name: the name of the configuration project.
    :param block_dev: the name of block device.
    :param network_dev: the name of network device.
    :param test: whether test the commands in file or not.
    :return: None
    """
    if iterations <= 10:
        print('-t iterations must be > 10, the input is ', iterations)
        return False
    if not os.path.exists(in_dir):
        print("Failed: The input directory is not existed:", in_dir)
        return False
    if not os.path.exists(out_dir):
        print("Warning: The output directory is not existed:", out_dir)
        return False

    in_file_list = glob.glob((str(in_dir) + "*.xlsx"))
    if not in_file_list:
        print("No .xlsx files exist in the directory")
        return False

    for file in in_file_list:
        in_file_name = file
        out_file_name = file.replace(".xlsx", ".yaml")
        if os.path.exists(str(out_dir) + out_file_name):
            print("Warning: The output yaml file is already exist, overwrite it!--", out_file_name)

        xlsx2yaml = XLSX2YAML((in_dir + in_file_name), (out_dir + out_file_name), project_name, iterations)
        result = xlsx2yaml.translate(block_dev, network_dev, test)
        if result:
            print('Translation of ' + str(file) + ' SUCCEEDED!\n')
        else:
            print('Translation of' + str(file) + ' FAILED!\n')

    return result


if __name__ == '__main__':
    ARG_PARSER = argparse.ArgumentParser(description="translate excel files to yaml files")
    ARG_PARSER.add_argument('-i', '--in_dir', metavar='INPUT DIRECTORY',
                            default="./", help='The folder of input excel files')
    ARG_PARSER.add_argument('-o', '--out_dir', metavar='OUTPUT DIRECTORY',
                            default="./", help='The folder of output yaml files')
    ARG_PARSER.add_argument('-t', '--iteration', metavar='ITERATIONS', type=int,
                            default="100", help='Iterations of the project (> 10)')
    ARG_PARSER.add_argument('-p', '--prj_name', metavar='NAME',
                            default="example", help='The name of the project')
    ARG_PARSER.add_argument('-bd', '--block_device', metavar='BLOCK DEVICE',
                            default="sda", help='The name of block device')
    ARG_PARSER.add_argument('-nd', '--network_device', metavar='NETWORK DEVICE',
                            default="enp189s0f0", help='The name of network device')
    ARG_PARSER.add_argument('-f', '--test', metavar='TEST COMMAND',
                            default="True", help='Whether test the command or not')

    ARGS = ARG_PARSER.parse_args()
    main(ARGS.in_dir, ARGS.out_dir, ARGS.iteration, ARGS.prj_name, ARGS.block_device, ARGS.network_device, ARGS.test)
